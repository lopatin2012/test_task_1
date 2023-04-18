from openpyxl import load_workbook
from openpyxl.styles import PatternFill, colors

# Переменные месторождений
#----------------------------------------------------------------------------------------------------------------------------------------------------
lst_field = ["М/р Серебристо-Зеленое", "М/р ВидноеЗолотистое", "М/р Шрамское", " М/р Минеральное",
                       "В Штрейдское", "М/р КрасноРусскинское", " М/р КрасноРусскинское", "М/р Ягарское",
                       "М/р Зап-ШтрейДское", "М/р ШтрейДское", "М/р Штрамское", "М/р Богучарское",
                       "М/р Власовское"]
# Переменные стиля ячейки. Заливка
#----------------------------------------------------------------------------------------------------------------------------------------------------
red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
green_fill = PatternFill(start_color='7fe757', end_color='7fe757', fill_type='solid')
yellow_fill = PatternFill(start_color='FFFF66', end_color='FFFF66', fill_type='solid')





# Шахматка. Колонки для выборки: B - месторождение, E - скважина, F - пласт
# TRDS. Колонки для выборки: F - месторождения, H - скважина, L - пласт/объект разработки
# Fond_ESP. Колонки для выборки: D - месторождение, G - скважина

# Создание словаря
#----------------------------------------------------------------------------------------------------------------------------------------------------
def load_dict_excel_shamatka(excel=''):
    number_excel = 1
    dict_excel = dict()
    for _ in range(280):
        for row in excel[f'A{number_excel}': f'ED{number_excel}']:
            for cell in row:
                if str(cell.value) in lst_field:
                    dict_excel[f'A{number_excel}'] = [
                        excel[f'B{number_excel}'].value, excel[f'E{number_excel}'].value, excel[f'F{number_excel}'].value
                        ]
        number_excel += 1
    return dict_excel


def load_dict_excel_trds(excel=''):
    number_excel = 1
    dict_excel = dict()
    for _ in range(280):
        for row in excel[f'A{number_excel}': f'AE{number_excel}']:
            for cell in row:
                if str(cell.value) in lst_field:
                    if str(excel[f'F{number_excel}'].value) != 'None':
                        dict_excel[f'A{number_excel}'] = [
                            excel[f'F{number_excel}'].value, excel[f'H{number_excel}'].value, excel[f'L{number_excel}'].value
                            ]
        number_excel += 1
    return dict_excel


def load_dict_excel_fond_esp(excel=''):
    number_excel = 1
    dict_excel = dict()
    for _ in range(280):
        for row in excel[f'A{number_excel}': f'J{number_excel}']:
            for cell in row:
                if str(cell.value) in lst_field:
                    dict_excel[f'A{number_excel}'] = [
                        excel[f'D{number_excel}'].value, excel[f'G{number_excel}'].value
                        ]
        number_excel += 1
    return dict_excel


def load_dict_excel_report(excel=''):
    number_excel = 1
    dict_excel = dict()
    for _ in range(70):
        for row in excel[f'A{number_excel}': f'ED{number_excel}']:
            for cell in row:
                if str(cell.value) in lst_field:
                    dict_excel[f'A{number_excel}'] = [
                        excel[f'H{number_excel}'].value, excel[f'I{number_excel}'].value, excel[f'AA{number_excel}'].value
                        ]
        number_excel += 1
    return dict_excel


# Задание №1. Проверка в отчёте. Если ни в одном другом файле нет данных, то подсвечивать строку красным цветом(Bad)
#----------------------------------------------------------------------------------------------------------------------------------------------------
def test_1(excel='', dict_excel=dict(), set=dict()):
    lst_set = []
    number = 0
    # создаём один список всех значений. В указанном задании важно наличие хотя бы одного значения в списке данных
    for _ in range(len(set)):
        for i in set[number].values():
            if ' '.join(map(str, i)) not in lst_set:
                lst_set.append(' '.join(map(str, i)))
        number += 1
    number = 15
    for value in dict_excel.values():
        if ' '.join(map(str, value)) not in lst_set:
            # Подсветка красным
            for row in excel[f'A{number}': f'EC{number}']:
                for cell in row:
                    cell.fill = red_fill
        # else:
        #     # Подсветка зелёным
        #     for row in excel[f'A{number}': f'EC{number}']:
        #         for cell in row:
        #             cell.fill = green_fill
        number += 1
    

# Задание №2. Проверка в других отчётах. Если в экселе отсутствует строка из отчёта, то подсвечивать строку в красный цвет(Bad)
#----------------------------------------------------------------------------------------------------------------------------------------------------
def test_2_shamatka(excel='', dict_excel=dict(), dict_report=dict()):
    lst_excel = []
    lst_report = []
    number = 1
    for i in dict_excel.values():
        if ' '.join(map(str, i)) not in lst_excel:
            lst_excel.append(' '.join(map(str, i)))
    for i in dict_report.values():
        if ' '.join(map(str, i)) not in lst_report:
            lst_report.append(' '.join(map(str, i)))
    for _ in range(280):
        for row in excel[f'B{number}': f'AD{number}']:
            for cell in row:
                if str(cell.value) in lst_field:
                    if f"{excel['B' + str(number)].value} {excel['E' + str(number)].value} {excel['F' + str(number)].value}" not in lst_report:
                        for row in excel[f'B{number}': f'R{number}']:
                            for cell in row:
                                cell.fill = red_fill
        number += 1


def test_2_trds(excel='', dict_excel=dict(), dict_report=dict()):
    lst_excel = []
    lst_report = []
    number = 1
    for i in dict_excel.values():
        if ' '.join(map(str, i)) not in lst_excel:
            lst_excel.append(' '.join(map(str, i)))
    for i in dict_report.values():
        if ' '.join(map(str, i)) not in lst_report:
            lst_report.append(' '.join(map(str, i)))
    for _ in range(280):
        for row in excel[f'B{number}': f'AD{number}']:
            for cell in row:
                if str(cell.value) in lst_field:
                    if f"{excel['F' + str(number)].value} {excel['H' + str(number)].value} {excel['L' + str(number)].value}" not in lst_report:
                        for row in excel[f'A{number}': f'AD{number}']:
                            for cell in row:
                                cell.fill = red_fill
        number += 1


def test_2_fond_esp(excel='', dict_excel=dict(), dict_report=dict()):
    lst_excel = []
    lst_report = []
    number = 1
    for i in dict_excel.values():
        if ' '.join(map(str, i)) not in lst_excel:
            lst_excel.append(' '.join(map(str, i)))
    for i in dict_report.values():
        if ' '.join(map(str, i)) not in lst_report:
            lst_report.append(' '.join(map(str, i[:2])))
    for _ in range(280):
        for row in excel[f'A{number}': f'J{number}']:
            for cell in row:
                if str(cell.value) in lst_field:
                    if f"{excel['D' + str(number)].value} {excel['G' + str(number)].value}" not in lst_report:
                        for row in excel[f'A{number}': f'J{number}']:
                            for cell in row:
                                cell.fill = red_fill
        number += 1



# Задание №3. Получаем необходимые значения в отчёте
#----------------------------------------------------------------------------------------------------------------------------------------------------

def load_dict_excel_report_test3_shamatka_1(excel=''):
    number_excel = 1
    dict_excel = dict()
    for _ in range(70):
        for row in excel[f'A{number_excel}': f'ED{number_excel}']:
            for cell in row:
                if str(cell.value) in lst_field:
                    dict_excel[f'A{number_excel}'] = [
                        excel[f'H{number_excel}'].value, excel[f'U{number_excel}'].value, excel[f'S{number_excel}'].value
                    ]
        number_excel += 1
    return dict_excel


def load_dict_excel_report_test3_shamatka_2(excel=''):
    number_excel = 1
    dict_excel = dict()
    for _ in range(70):
        for row in excel[f'A{number_excel}': f'ED{number_excel}']:
            for cell in row:
                if str(cell.value) in lst_field and str(excel[f'AO{number_excel}'].value) != str(None):
                    dict_excel[f'A{number_excel}'] = [
                        excel[f'H{number_excel}'].value, excel[f'AO{number_excel}'].value
                    ]
        number_excel += 1
    return dict_excel


def load_dict_excel_report_test3_trds(excel=''):
    number_excel = 1
    dict_excel = dict()
    for _ in range(70):
        for row in excel[f'A{number_excel}': f'ED{number_excel}']:
            for cell in row:
                if str(cell.value) in lst_field:
                    dict_excel[f'A{number_excel}'] = [
                        f"{str(excel[f'H{number_excel}'].value)} {str(excel[f'AD{number_excel}'].value)}"
                    ]
        number_excel += 1
    return dict_excel


def load_dict_excel_report_test3_fond_esp(excel=''):
    number_excel = 1
    dict_excel = dict()
    for _ in range(70):
        for row in excel[f'A{number_excel}': f'ED{number_excel}']:
            for cell in row:
                if str(cell.value) in lst_field:
                    dict_excel[f'A{number_excel}'] = [
                        f"{str(excel[f'H{number_excel}'].value)}, {str(excel[f'AE{number_excel}'].value)}"
                    ]
        number_excel += 1
    return dict_excel


def load_dict_excel_shamatka_1_test3(excel=''):
    number_excel = 1
    dict_excel = dict()
    for _ in range(280):
        for row in excel[f'A{number_excel}': f'ED{number_excel}']:
            for cell in row:
                if str(cell.value) in lst_field:
                    dict_excel[f'A{number_excel}'] = [
                        excel[f'B{number_excel}'].value, excel[f'G{number_excel}'].value, excel[f'M{number_excel}'].value
                    ]
        number_excel += 1
    return dict_excel


def load_dict_excel_shamatka_2_test3(excel=''):
    number_excel = 1
    dict_excel = dict()
    for _ in range(280):
        for row in excel[f'A{number_excel}': f'ED{number_excel}']:
            for cell in row:
                if str(cell.value) in lst_field:
                    dict_excel[f'A{number_excel}'] = [
                        excel[f'B{number_excel}'].value, excel[f'G{number_excel}'].value, excel[f'O{number_excel}'].value
                    ]
        number_excel += 1
    return dict_excel


def load_dict_excel_trds_test3(excel=''):
    number_excel = 1
    dict_excel = dict()
    for _ in range(280):
        for row in excel[f'A{number_excel}': f'ED{number_excel}']:
            for cell in row:
                if str(cell.value) in lst_field:
                    dict_excel[f'A{number_excel}'] = [
                        excel[f'B{number_excel}'].value, excel[f'W{number_excel}'].value
                    ]
        number_excel += 1
    return dict_excel



def load_dict_excel_report_1_test_3(excel=''):
    number_excel = 1
    dict_excel = dict()
    for _ in range(280):
        for row in excel[f'A{number_excel}': f'ED{number_excel}']:
            for cell in row:
                if str(cell.value) in lst_field:
                    dict_excel[f'A{number_excel}'] = [
                        excel[f'B{number_excel}'].value, excel[f'E{number_excel}'].value, excel[f'F{number_excel}'].value
                        ]
        number_excel += 1
    return dict_excel



def load_dict_excel_fond_esp_test3(excel=''):
    number_excel = 1
    dict_excel = dict()
    for _ in range(280):
        for row in excel[f'A{number_excel}': f'ED{number_excel}']:
            for cell in row:
                if str(cell.value) in lst_field:
                    dict_excel[f'A{number_excel}'] = [
                        excel[f'B{number_excel}'].value, excel[f'H{number_excel}'].value
                    ]
        number_excel += 1
    return dict_excel

# Функции для теста №3
# Реализация без поиска значений. Проверка работоспособности
#----------------------------------------------------------------------------------------------------------------------------------------------------
def last_day_test_3_shamatka_1(excel_1='', excel_2=''):
    if excel_1['M37'].value == excel_2['P17'].value:
        excel_1['M37'].fill = green_fill
        excel_2['P17'].fill = green_fill
    else:
        excel_1['M37'].fill = yellow_fill
        excel_2['P17'].fill = yellow_fill


def last_day_test_3_shamatka_2(excel_1='', excel_2=''):
    if excel_1['O8'].value == excel_2['AO15'].value:
        excel_1['O8'].fill = green_fill
        excel_2['AO15'].fill = green_fill
    else:
        excel_1['O8'].fill = yellow_fill
        excel_2['AO15'].fill = yellow_fill
        
# Упрощенные функции для проверки совпадения значения в ячейке
#----------------------------------------------------------------------------------------------------------------------------------------------------
def one_in_one_trds_test_3(excel_1='', excel_2=''):
    lst_trds = []
    lst_report = []

    for i in range(22, 52):
        if str(excel_1['F' + str(i)].value) + ' ' + str(excel_1['W' + str(i)].value) not in lst_trds:
            lst_trds.append(str(excel_1['F' + str(i)].value) + ' ' + str(excel_1['W' + str(i)].value))
            
    for j in range(15, 64):
        if str(excel_2['H' + str(j)].value) + ' ' + str(excel_2['AE' + str(j)].value) not in lst_report:
            lst_report.append(str(excel_2['H' + str(j)].value) + ' ' + str(excel_2['AE' + str(j)].value)) 
               
    for i in range(22, 52):
        if str(excel_1['F' + str(i)].value) + ' ' + str(excel_1['W' + str(i)].value) in lst_report:
            excel_1['W' + str(i)].fill = green_fill
        else:
            excel_1['W' + str(i)].fill = yellow_fill
            
    for j in range(15, 64):
        if str(excel_2['H' + str(j)].value) + ' ' + str(excel_2['AE' + str(j)].value) in lst_trds:
            excel_2['AE' + str(j)].fill = green_fill
        else:
            excel_2['AE' + str(j)].fill = yellow_fill
            
            
def one_in_one_fond_esp_test_3(excel_1='', excel_2=''):
    lst_fond_esp = []
    lst_report = []

    for i in range(2, 63):
        if str(excel_1['D' + str(i)].value) + ' ' + str(excel_1['H' + str(i)].value) not in lst_fond_esp:
            lst_fond_esp.append(str(excel_1['D' + str(i)].value) + ' ' + str(excel_1['H' + str(i)].value))
            
    for j in range(15, 64):
        if str(excel_2['H' + str(j)].value) + ' ' + str(excel_2['AD' + str(j)].value) not in lst_report:
            lst_report.append(str(excel_2['H' + str(j)].value) + ' ' + str(excel_2['AD' + str(j)].value)) 
               
    for i in range(2, 63):
        if str(excel_1['D' + str(i)].value) + ' ' + str(excel_1['H' + str(i)].value) in lst_report:
            excel_1['H' + str(i)].fill = green_fill
        else:
            excel_1['H' + str(i)].fill = yellow_fill
            
    for j in range(15, 64):
        if str(excel_2['H' + str(j)].value) + ' ' + str(excel_2['AD' + str(j)].value) in lst_fond_esp:
            excel_2['AD' + str(j)].fill = green_fill
        else:
            excel_2['AD' + str(j)].fill = yellow_fill


# Создаём ссылки на файлы
#----------------------------------------------------------------------------------------------------------------------------------------------------
excel_book_1 = load_workbook(filename="ШАХМАТКА 11.2022.xlsx")
excel_book_2 = load_workbook(filename="ШАХМАТКА 12.2022.xlsx")
excel_book_3 = load_workbook(filename="TRDS 12.2022.xlsx")
excel_book_4 = load_workbook(filename="Fond_ESP 12.2022.xlsx")
excel_book_5 = load_workbook(filename="ОТЧЕТ 01.2023.xlsx")

# Выбор активной страницы для работы с данными
# Ранее была попытка реализовать это через функцию, однако не сохранялись изменения в файле. Нужно разобраться
excel_book_active_1 = excel_book_1.active
excel_book_active_2 = excel_book_2.active
excel_book_active_3 = excel_book_3.active
excel_book_active_4 = excel_book_4.active
excel_book_active_5 = excel_book_5.active


# Создаём словари с необходимыми данными
#----------------------------------------------------------------------------------------------------------------------------------------------------
dict_excel_shamatka_1 = load_dict_excel_shamatka(excel=excel_book_active_1)
dict_excel_shamatka_2 = load_dict_excel_shamatka(excel=excel_book_active_2)
dict_excel_trds = load_dict_excel_trds(excel=excel_book_active_3)
dict_excel_fond_esp = load_dict_excel_fond_esp(excel=excel_book_active_4)
dict_excel_report = load_dict_excel_report(excel=excel_book_active_5)

# Создаём первичные словари отчётов для теста №3
#----------------------------------------------------------------------------------------------------------------------------------------------------
dict_excel_report_test_3_shamatka_1 = load_dict_excel_report_test3_shamatka_1(excel=excel_book_active_5)
dict_excel_report_test_3_shamatka_2 = load_dict_excel_report_test3_shamatka_2(excel=excel_book_active_5)
dict_excel_report_test_3_trds = load_dict_excel_report_test3_trds(excel=excel_book_active_5)
dict_excel_report_test_3_fond_esp = load_dict_excel_report_test3_fond_esp(excel=excel_book_active_5)
# Создаём вторичные словари отчётов для теста №3
#----------------------------------------------------------------------------------------------------------------------------------------------------
dict_excel_test_3_shamatka_1 = load_dict_excel_shamatka_1_test3(excel=excel_book_active_1)
dict_excel_test_3_shamatka_2 = load_dict_excel_shamatka_2_test3(excel=excel_book_active_2)
dict_excel_test_3_trds = load_dict_excel_trds_test3(excel=excel_book_active_3)
dict_excel_test_3_fond_esp = load_dict_excel_fond_esp_test3(excel=excel_book_active_4)

# Множество из словарей
#----------------------------------------------------------------------------------------------------------------------------------------------------
dict_excel_set = dict_excel_shamatka_1, dict_excel_shamatka_2, dict_excel_trds, dict_excel_fond_esp

# Запуск тестового задания №1
#----------------------------------------------------------------------------------------------------------------------------------------------------
test_1(excel=excel_book_active_5,
       dict_excel=dict_excel_report, set=dict_excel_set)

# Запуск тестового задания №2

test_2_shamatka(excel=excel_book_active_1, dict_excel=dict_excel_shamatka_1,
       dict_report=dict_excel_report)

test_2_shamatka(excel=excel_book_active_2, dict_excel=dict_excel_shamatka_2,
                dict_report=dict_excel_report)

test_2_trds(excel=excel_book_active_3, dict_excel=dict_excel_trds,
                dict_report=dict_excel_report)

test_2_fond_esp(excel=excel_book_active_4, dict_excel=dict_excel_fond_esp,
            dict_report=dict_excel_report)

# Запуск тестового заданий №3
last_day_test_3_shamatka_1(excel_1=excel_book_active_2, excel_2=excel_book_active_5)
last_day_test_3_shamatka_2(excel_1=excel_book_active_1, excel_2=excel_book_active_5)
one_in_one_trds_test_3(excel_1=excel_book_active_3, excel_2=excel_book_active_5)
one_in_one_fond_esp_test_3(excel_1=excel_book_active_4, excel_2=excel_book_active_5)


excel_book_1.save(filename='Шахматка 11 test1.xlsx')
excel_book_2.save(filename='Шахматка 12 test2.xlsx')
excel_book_3.save(filename='TRDS test3.xlsx')
excel_book_4.save(filename='Fond_ESP test4.xlsx')
excel_book_5.save(filename='ОТЧЕТ test5.xlsx')